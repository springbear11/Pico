from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT = Path(__file__).with_name("PicoATE_TaskEngine_Development_Plan.xlsx")
BASELINE_START = date(2026, 6, 29)
HOURS_PER_DAY = 8

PHASES = [
    ("P1", "架构与工程骨架", "确定三层边界并建立可独立测试的任务引擎工程"),
    ("P2", "调度运行时内核", "完成图调度、多 UUT、资源、Barrier、错误策略和 Cleanup"),
    ("P3", "流程模型与编译器", "实现 JSON 到编辑模型再到不可变执行计划"),
    ("P4", "结果、报告与 CLI", "提供不暴露调度内部状态的结果 DTO 和命令行验证入口"),
    ("P5", "业务模块与跨进程边界", "让业务逻辑通过模块接口和 Transport 接入，不修改调度器"),
    ("P6", "循环与变量能力", "实现调度层循环、配置变量和运行期变量替换"),
    ("P7", "设备生命周期与适配验证", "按逻辑设备 ID 管理长连接，并验证 DMM/CAN 适配方向"),
    ("P8", "质量保障与文档", "用测试、示例和文档固化架构边界与交付方式"),
]

# 标准工时是一个有 C++/Qt 经验的工程师完成原型级实现、单元测试和必要文档的估算。
TASKS = [
    ("P1.1", "P1", "架构边界与 V3.1 文档", "明确 UI、任务调度、业务逻辑三层；定义 ExecutionPlan、运行期控制器、扩展点和持久化预留。", 16, "高", "P0", "-", "架构文档、目录边界、关键接口轮廓", "文档评审通过", "≤2026-06-20", "PicoATE_Architecture_Design_Document.md"),
    ("P1.2", "P1", "VS2022 + Qt6 + CMake 工程骨架", "建立 Core、Tests、CLI 等目标，支持 VS2022 Preset 编译。", 8, "中", "P0", "P1.1", "可生成并编译的解决方案", "Debug 构建通过", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P1.3", "P1", "公共类型与不可变 ExecutionPlan", "定义节点、边、CleanupRegion、Frame、Activation、Attempt 等核心值类型。", 16, "高", "P0", "P1.2", "ExecutionPlan 与 Runtime DTO", "CoreTests 类型与引用校验", "≤2026-06-20", "progress_plan_2026-06-20.md"),

    ("P2.1", "P2", "ResourceManager", "实现资源互斥、租约、等待队列及快照中的 waiter 信息。", 16, "高", "P0", "P1.3", "资源申请/释放/等待 API", "资源竞争与释放测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P2.2", "P2", "BarrierController", "实现多 UUT 到达、释放、失败成员移除、超时策略及快照。", 20, "高", "P0", "P1.3", "Barrier 独立控制器", "多 UUT Barrier 集成测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P2.3", "P2", "ErrorPolicyEngine", "实现 Fail/Error/Timeout 的 Continue、Stop、Retry、Cleanup、Abort 决策。", 12, "中", "P0", "P1.3", "错误策略决策链", "错误路径测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P2.4", "P2", "ExecutionGraphScheduler 主循环", "实现 ready 判定、节点启动、完成回收、依赖推进和终态判断。", 24, "很高", "P0", "P2.1,P2.2,P2.3", "同步事件驱动调度器", "基础图执行测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P2.5", "P2", "Retry/Cleanup/alwaysRun/Stop 语义", "补齐重试 attempt、Finally Cleanup、停止后清理以及 alwaysRun 约束。", 20, "很高", "P0", "P2.4", "完整异常与停止路径", "失败清理、重试、停止测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P2.6", "P2", "ExecutionSession 多 UUT 编排", "统一持有共享调度组件，支持多 UUT 运行、requestStop 和会话快照骨架。", 16, "高", "P0", "P2.4", "ExecutionSession", "多 UUT 端到端测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P2.7", "P2", "调度集成测试强化", "覆盖 Barrier 释放、失败成员移除、资源等待、Retry、Cleanup 和 Stop。", 16, "高", "P0", "P2.5,P2.6", "调度回归测试集", "CTest 通过", "≤2026-06-20", "progress_plan_2026-06-20.md"),

    ("P3.1", "P3", "SequenceDef 编辑模型", "定义 Sequence、Group、Step、资源、Retry、Timeout、Error、Barrier 等配置模型。", 12, "中", "P0", "P1.3", "SequenceDef/StepDef/GroupDef", "模型转换与重复 ID 测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P3.2", "P3", "PlanBuilder", "把编辑模型编译为 Node、Edge、Finally 边和 CleanupRegion。", 24, "很高", "P0", "P3.1,P2.4", "SequenceDef -> ExecutionPlan", "三层贯通测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P3.3", "P3", "SequenceCompiler JSON 解析", "实现 JSON -> SequenceDef -> ExecutionPlan，并提供精确字段路径错误。", 24, "高", "P0", "P3.1,P3.2", "SequenceCompiler", "JSON 编译执行测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P3.4", "P3", "类型校验与 Warning 机制", "区分缺省值和类型错误；未知字段给 warning，x-* 和 vendor 作为扩展口。", 16, "高", "P1", "P3.3", "CompileError/CompileWarning", "坏 JSON 一次捕获多项错误", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P3.5", "P3", "Disabled/Custom/Checkpoint 语义", "支持禁用组和步骤、自定义组桥接、checkpointBefore/After 映射。", 12, "中", "P1", "P3.2,P3.3", "扩展后的编译语义", "示例与编译测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P3.6", "P3", "Sequence Schema 与示例", "提供 simple/basic/custom 等 JSON 示例和字段说明。", 8, "低", "P1", "P3.3", "Schema 文档与示例", "CLI 可加载示例", "≤2026-06-20", "progress_plan_2026-06-20.md"),

    ("P4.1", "P4", "PicoATE.Cli", "支持读取 sequence、编译、运行多 UUT 并打印步骤和 attempt。", 12, "中", "P1", "P3.3,P2.6", "命令行 Runner", "CLI CTest", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P4.2", "P4", "ExecutionReport 只读 DTO", "把运行状态复制为面向 CLI、报表和 UI 的稳定值对象。", 12, "中", "P0", "P2.6", "ExecutionReport/UutReport/StepReport", "报告排序与错误标记测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P4.3", "P4", "Measurement 与 Limit 模型", "正式表达测量名称、值、单位、上下限、判定和原始数据。", 12, "中", "P0", "P4.2,P5.1", "MeasurementResult", "Module -> Node -> Report -> CLI 测试", "2026-06-24", "work_summary_2026-06-24.md"),
    ("P4.4", "P4", "Loop Report iteration 元数据", "报告显式给出循环归属、迭代序号和值，避免 UI 自行猜测。", 8, "中", "P1", "P4.2,P6.1", "Loop iteration report", "循环报告测试", "2026-06-25", "progress_plan_2026-06-20.md"),

    ("P5.1", "P5", "IModule 与 ModuleRegistry", "业务 Action 通过模块接口执行，ModuleResult 再由运行器转换为 NodeResult。", 16, "高", "P0", "P2.4", "模块运行时边界", "Fake module 映射测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P5.2", "P5", "IModuleTransport 与 JSON DTO", "建立跨进程/跨语言请求响应协议，并集中实现 JSON 序列化。", 12, "高", "P0", "P5.1", "Transport contract", "JSON 往返测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P5.3", "P5", "QProcessTransport 与 MockHost", "通过 stdio 一行 JSON 调用独立进程，验证正常、超时和崩溃处理。", 24, "很高", "P0", "P5.2", "QProcessTransport/MockHost", "跨进程端到端测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P5.4", "P5", "moduleBindings 与 Python 示例", "从 JSON 自动注册外部模块，并验证 Python 脚本调用。", 16, "高", "P1", "P5.3,P3.3", "配置化外部模块接入", "Python echo CLI/Core 测试", "≤2026-06-20", "progress_plan_2026-06-20.md"),
    ("P5.5", "P5", "DllBridgeInvoker 与测试 DLL", "进程内验证 QLibrary、统一 C ABI、缓冲区和超时报告。", 20, "很高", "P0", "P5.2", "DLL Bridge 原型", "成功/错误/超时测试", "2026-06-21", "progress_plan_2026-06-20.md"),
    ("P5.6", "P5", "NativeHost 独立进程", "在子进程加载 DLL，隔离厂商 DLL 崩溃或卡死对主程序的影响。", 24, "很高", "P0", "P5.3,P5.5", "PicoATE.NativeHost.exe", "NativeHost 端到端测试", "2026-06-21", "progress_plan_2026-06-20.md"),
    ("P5.7", "P5", "NativeHost manifest", "配置 DLL 路径、symbol、buffer、timeout 和 metadata，支持变量注入。", 16, "高", "P1", "P5.6,P6.2", "Manifest parser/CLI", "manifest sequence 测试", "2026-06-22", "work_summary_2026-06-22.md"),
    ("P5.8", "P5", "纯软件 CAN DLL 示例", "在无 CAN 分析仪条件下验证报文解码、缩放、限值和错误结果。", 16, "高", "P1", "P5.6,P5.7,P4.3", "CAN example DLL/manifest/sequence", "Pass/Fail CLI/Core 测试", "2026-06-24", "work_summary_2026-06-24.md"),
    ("P5.9", "P5", "外部模块协议文档", "明确 C/C++ DLL、Python 和通用 exe 的接入边界与请求响应格式。", 8, "低", "P1", "P5.2", "module_contract.md", "文档与示例一致", "≤2026-06-20", "progress_plan_2026-06-20.md"),

    ("P6.1", "P6", "LoopController For 循环", "循环游标由调度层维护，Plan 保持不可变，业务模块对循环无感。", 24, "很高", "P0", "P2.4,P3.2", "LoopController/LoopRegion", "循环 JSON 与 CLI/Core 测试", "2026-06-22", "work_summary_2026-06-22.md"),
    ("P6.2", "P6", "VariableResolver 中心化", "支持内置变量、显式变量、环境变量、递归 Map/List 和路径化错误。", 16, "高", "P0", "P3.3", "公共 VariableResolver", "递归替换测试", "2026-06-22", "work_summary_2026-06-22.md"),
    ("P6.3", "P6", "运行期与 Loop 变量替换", "节点执行前解析 payload/inputs，支持 loop、UUT、attempt 和 whole-field 类型。", 16, "高", "P0", "P6.1,P6.2,P5.1", "RuntimeVariableResolver", "每轮不同输入测试", "2026-06-24", "work_summary_2026-06-24.md"),

    ("P7.1", "P7", "硬件生命周期设计与 DeviceSessionManager", "定义逻辑设备、工位级 session、连接复用、closeAll 和错误模型。", 24, "很高", "P0", "P5.1", "DeviceSessionManager", "Fake session 生命周期测试", "2026-06-25", "hardware_module_lifecycle_design.md"),
    ("P7.2", "P7", "Station config JSON", "把 DMM1/CAN1 的 driver、address、lifetime 和 options 从 sequence 中拆出。", 16, "高", "P0", "P7.1,P6.2", "StationConfig parser", "字段校验与变量替换测试", "2026-06-26", "work_summary_2026-06-26.md"),
    ("P7.3", "P7", "StationRuntime 与 CLI --station", "统一初始化工位设备并在运行前输出设备摘要。", 12, "中", "P1", "P7.2,P4.1", "StationRuntime/CLI integration", "Station 示例 CLI 测试", "2026-06-26", "work_summary_2026-06-26.md"),
    ("P7.4", "P7", "PersistentQProcessTransport", "长驻 Host 跨 Step 保持设备状态，支持 health/status/reconnect/shutdown。", 24, "很高", "P0", "P5.3,P7.1", "Persistent transport/FakeInstrumentHost", "跨 Step 状态与恢复测试", "2026-06-26", "work_summary_2026-06-26.md"),
    ("P7.5", "P7", "业务模块设备服务与 Proxy", "ModuleExecutionContext 按 DMM1/CAN1 获取设备代理，不直接 new driver。", 16, "高", "P0", "P7.1,P7.4", "IModuleRuntimeServices/TransportDeviceSession", "模块到设备 session 测试", "2026-06-27", "progress_plan_2026-06-20.md"),
    ("P7.6", "P7", "DMM/CAN Adapter spike", "用 Fake Host 验证 Connect/Configure/Read/Disconnect 与 CAN ReadFrame 适配方式。", 24, "很高", "P0", "P7.5,P4.3", "example.dmm/example.can", "JSON -> Session -> Adapter E2E", "2026-06-27", "progress_plan_2026-06-20.md"),

    ("P8.1", "P8", "自动化构建与回归测试", "持续维护 Core、CLI、Host 和示例的 CTest 验证。", 12, "高", "P0", "所有实现任务", "VS2022 Debug build + CTest", "12/12 测试通过", "2026-06-27", "progress_plan_2026-06-20.md"),
    ("P8.2", "P8", "架构、能力与进度文档", "更新项目愿景、模块运行时、设备生命周期、当前能力拆解和计划。", 12, "中", "P1", "所有实现任务", "docs 文档集", "代码、示例、文档交叉检查", "2026-06-28", "docs/"),
]

FUTURE_TASKS = [
    ("N1", "Engine/UI 工程物理拆分", "把任务引擎和 Qt Runner 做成独立 CMake target；顶层可一起编译，也可单独编译。", 16, "高", "阻塞 UI 起步", "未开始"),
    ("N2", "ExecutionViewModel 与异步 Runner", "UI 只通过命令和只读 DTO 使用引擎，执行放到工作线程避免卡界面。", 32, "很高", "阻塞 UI Runner", "未开始"),
    ("N3", "Host/Transport 诊断增强", "补齐 stderr/stdout 截断、启动失败、协议错误、timeout 和崩溃诊断。", 16, "高", "不阻塞 UI", "待办"),
    ("N4", "Session Checkpoint 持久化", "实现 ISessionPersistence、资源 waiter 恢复和运行会话恢复策略。", 40, "很高", "不阻塞第一版 UI", "架构预留"),
    ("N5", "暂停/继续与调试会话", "实现安全暂停点、继续执行、断点和调试状态恢复。", 32, "很高", "不阻塞第一版 UI", "待办"),
    ("N6", "并发与压力测试", "验证大量 UUT、长序列、共享 Persistent Transport 和资源竞争。", 32, "很高", "发布前需要", "待办"),
    ("N7", "真实 DMM/CAN 硬件验证", "用厂商驱动和真实设备验证连接、重连、超时、报文和清理。", 40, "很高", "依赖硬件", "受阻于设备条件"),
    ("N8", "Engine SDK/export/package", "增加 install/export 和 find_package，使 UI 或第三方程序可脱离源码树引用引擎。", 24, "高", "发布前需要", "待办"),
]


NAVY = "17365D"
BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
TEAL = "2A7F77"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
AMBER = "FFC000"
LIGHT_AMBER = "FFF2CC"
RED = "C00000"
LIGHT_RED = "FCE4D6"
GRAY = "7F8C8D"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"

thin_gray = Side(style="thin", color="D9E1F2")
border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)


def is_business_day(value: date) -> bool:
    return value.weekday() < 5


def next_business_day(value: date) -> date:
    value += timedelta(days=1)
    while not is_business_day(value):
        value += timedelta(days=1)
    return value


def build_schedule(tasks):
    current = BASELINE_START
    used_half_days = 0
    result = []
    for task in tasks:
        slots = task[4] // 4
        start = current
        remaining = slots
        end = current
        while remaining > 0:
            available = 2 - used_half_days
            consumed = min(available, remaining)
            remaining -= consumed
            used_half_days += consumed
            end = current
            if used_half_days == 2 and remaining > 0:
                current = next_business_day(current)
                used_half_days = 0
        if used_half_days == 2:
            current = next_business_day(current)
            used_half_days = 0
        result.append((task, start, end))
    return result


def style_title(ws, title: str, subtitle: str, end_column: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, title)
    cell.font = Font(name="Microsoft YaHei", size=18, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    cell = ws.cell(2, 1, subtitle)
    cell.font = Font(name="Microsoft YaHei", size=10, color="44546A")
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30


def style_header(ws, row: int, start_column: int, end_column: int):
    for col in range(start_column, end_column + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="Microsoft YaHei", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 28


def style_body(ws, start_row: int, end_row: int, start_column: int, end_column: int):
    for row in range(start_row, end_row + 1):
        fill = PatternFill("solid", fgColor=WHITE if row % 2 else "F8FBFD")
        for col in range(start_column, end_column + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Microsoft YaHei", size=9, color="1F1F1F")
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_overview(wb: Workbook, schedule):
    ws = wb.active
    ws.title = "项目概览"
    style_title(ws, "PicoATE 任务引擎研发计划与工作量评估", "任务引擎 WBS、标准工时、阶段排期、交付物与验收计划。", 8)

    summary = [
        ("项目目标", "构建 UI、任务调度、业务测试逻辑三层解耦的 ATE 平台内核；项目交付优先通过 JSON/config 和新增底层业务模块完成。"),
        ("本表范围", "任务调度引擎、流程编译、模块运行时、DLL/Python/进程边界、循环变量、设备生命周期、DMM/CAN 软件适配验证；不含正式 Qt UI。"),
        ("估算口径", "1 名具备 C++/Qt/CMake 经验的工程师；8 小时/人日；包含原型级实现、单元/集成测试和必要文档；不含采购、真实硬件联调等待和生产认证。"),
        ("计划周期", f"{schedule[0][1]:%Y-%m-%d} 至 {schedule[-1][2]:%Y-%m-%d}，按工作日连续排布。"),
        ("当前状态", "计划于 2026-06-29 启动，当前阶段为任务分解与排期确认。"),
        ("计划边界", "DMM/CAN 阶段先使用 Fake Host 和纯软件 DLL 验证接口与链路；真实硬件稳定性验证作为后续专项安排。"),
        ("阶段出口", "任务引擎计划完成后进行 Engine/UI 工程物理拆分，再进入 ExecutionViewModel 和最小 Qt Runner UI。"),
    ]

    row = 4
    for key, value in summary:
        ws.cell(row, 1, key)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.cell(row, 2, value)
        ws.cell(row, 1).font = Font(name="Microsoft YaHei", bold=True, color=WHITE)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=TEAL)
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row, 2).font = Font(name="Microsoft YaHei", size=10)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor="F3F8F7")
        ws.cell(row, 2).alignment = Alignment(vertical="center", wrap_text=True)
        for col in range(1, 9):
            ws.cell(row, col).border = border
        ws.row_dimensions[row].height = 42
        row += 1

    row += 1
    ws.cell(row, 1, "关键指标")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row, 1).font = Font(name="Microsoft YaHei", size=12, bold=True, color=WHITE)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row, 1).alignment = Alignment(vertical="center")
    row += 1
    metrics = [
        ("已拆解任务", len(TASKS), "项"),
        ("标准总工时", sum(item[4] for item in TASKS), "小时"),
        ("标准总人日", sum(item[4] for item in TASKS) / HOURS_PER_DAY, "人日"),
        ("完成阶段", len(PHASES), "个"),
        ("自动化测试", "12+", "项目标"),
        ("后续增强", len(FUTURE_TASKS), "项"),
    ]
    for index, (label, value, unit) in enumerate(metrics):
        col = 1 + (index % 3) * 2
        current_row = row + index // 3 * 2
        ws.cell(current_row, col, label)
        ws.cell(current_row, col + 1, value)
        ws.cell(current_row + 1, col + 1, unit)
        ws.cell(current_row, col).font = Font(name="Microsoft YaHei", bold=True, color="44546A")
        ws.cell(current_row, col + 1).font = Font(name="Microsoft YaHei", size=16, bold=True, color=BLUE)
        ws.cell(current_row + 1, col + 1).font = Font(name="Microsoft YaHei", size=9, color=GRAY)
        for r in (current_row, current_row + 1):
            for c in (col, col + 1):
                ws.cell(r, c).fill = PatternFill("solid", fgColor="F8FBFD")
                ws.cell(r, c).border = border
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    ws.row_dimensions[row + 1].height = 22
    ws.row_dimensions[row + 2].height = 28
    ws.row_dimensions[row + 3].height = 22

    note_row = row + 5
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 2, end_column=8)
    ws.cell(note_row, 1, "排期说明：工时按 1 名具备 C++/Qt/CMake 经验的工程师估算，按 8 小时/人日和周一至周五连续排布。Checkpoint 恢复、并发压力、真实硬件联调和 SDK 发布作为后续增强项，不计入任务引擎首轮开发基线。")
    ws.cell(note_row, 1).font = Font(name="Microsoft YaHei", size=10, bold=True, color="7F6000")
    ws.cell(note_row, 1).fill = PatternFill("solid", fgColor=LIGHT_AMBER)
    ws.cell(note_row, 1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.cell(note_row, 1).border = border

    set_widths(ws, {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18})
    ws.sheet_view.showGridLines = False


def add_task_plan(wb: Workbook, schedule):
    ws = wb.create_sheet("任务计划")
    style_title(ws, "任务引擎详细工作计划", "所有任务自 2026-06-29 起，按单人标准工时和前置依赖顺排。", 16)
    headers = ["WBS", "阶段", "工作项", "大白话说明", "难度", "优先级", "估算工时(h)", "估算人日", "计划开始", "计划结束", "状态", "是否阻塞 UI", "前置依赖", "主要交付物", "验收/验证", "备注"]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    style_header(ws, 4, 1, len(headers))

    phase_names = {code: name for code, name, _ in PHASES}
    for row, (task, start, end) in enumerate(schedule, 5):
        wbs, phase, name, desc, hours, difficulty, priority, dependency, output, verify, actual, source = task
        values = [
            wbs, f"{phase} {phase_names[phase]}", name, desc, difficulty, priority, hours,
            hours / HOURS_PER_DAY, start, end, "未开始", "否", dependency, output, verify,
            "生产化事项见“后续增强”工作表",
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)

    end_row = 4 + len(schedule)
    style_body(ws, 5, end_row, 1, len(headers))
    for row in range(5, end_row + 1):
        ws.cell(row, 7).number_format = "0"
        ws.cell(row, 8).number_format = "0.0"
        ws.cell(row, 9).number_format = "yyyy-mm-dd"
        ws.cell(row, 10).number_format = "yyyy-mm-dd"
        for col in (1, 5, 6, 7, 8, 9, 10, 11, 12, 16):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 50

    total_row = end_row + 2
    ws.cell(total_row, 1, "合计")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    ws.cell(total_row, 7, f"=SUM(G5:G{end_row})")
    ws.cell(total_row, 8, f"=SUM(H5:H{end_row})")
    for col in range(1, 17):
        ws.cell(total_row, col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(total_row, col).font = Font(name="Microsoft YaHei", bold=True, color=WHITE)
        ws.cell(total_row, col).border = border
        ws.cell(total_row, col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(total_row, 8).number_format = "0.0"

    ws.auto_filter.ref = f"A4:P{end_row}"
    ws.freeze_panes = "D5"
    ws.sheet_view.showGridLines = False
    ws.conditional_formatting.add(f"E5:E{end_row}", FormulaRule(formula=["E5=\"很高\""], fill=PatternFill("solid", fgColor=LIGHT_RED)))
    ws.conditional_formatting.add(f"K5:K{end_row}", FormulaRule(formula=["K5=\"未开始\""], fill=PatternFill("solid", fgColor=LIGHT_AMBER)))
    set_widths(ws, {
        "A": 9, "B": 24, "C": 30, "D": 52, "E": 9, "F": 9, "G": 12, "H": 12,
        "I": 13, "J": 13, "K": 11, "L": 13, "M": 22, "N": 34, "O": 32, "P": 30,
    })
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_phase_summary(wb: Workbook, schedule):
    ws = wb.create_sheet("阶段汇总")
    style_title(ws, "阶段工作量汇总", "汇总各阶段目标、投入、周期和计划完成率。", 10)
    headers = ["阶段", "阶段名称", "阶段目标", "任务数", "标准工时(h)", "标准人日", "计划开始", "计划结束", "当前状态", "完成率"]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    style_header(ws, 4, 1, len(headers))

    for row, (code, name, goal) in enumerate(PHASES, 5):
        phase_rows = [(task, start, end) for task, start, end in schedule if task[1] == code]
        hours = sum(task[4] for task, _, _ in phase_rows)
        values = [code, name, goal, len(phase_rows), hours, hours / HOURS_PER_DAY,
                  min(start for _, start, _ in phase_rows), max(end for _, _, end in phase_rows), "未开始", 0]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)

    end_row = 4 + len(PHASES)
    style_body(ws, 5, end_row, 1, len(headers))
    for row in range(5, end_row + 1):
        ws.cell(row, 6).number_format = "0.0"
        ws.cell(row, 7).number_format = "yyyy-mm-dd"
        ws.cell(row, 8).number_format = "yyyy-mm-dd"
        ws.cell(row, 10).number_format = "0%"
        ws.row_dimensions[row].height = 46
        for col in (1, 4, 5, 6, 7, 8, 9, 10):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    total_row = end_row + 2
    values = ["TOTAL", "任务引擎原型", "计划完成从 JSON 流程到模块/设备适配的可执行闭环", f"=SUM(D5:D{end_row})", f"=SUM(E5:E{end_row})", f"=SUM(F5:F{end_row})", schedule[0][1], schedule[-1][2], "未开始", 0]
    for col, value in enumerate(values, 1):
        ws.cell(total_row, col, value)
        ws.cell(total_row, col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(total_row, col).font = Font(name="Microsoft YaHei", bold=True, color=WHITE)
        ws.cell(total_row, col).border = border
        ws.cell(total_row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(total_row, 6).number_format = "0.0"
    ws.cell(total_row, 7).number_format = "yyyy-mm-dd"
    ws.cell(total_row, 8).number_format = "yyyy-mm-dd"
    ws.cell(total_row, 10).number_format = "0%"

    ws.conditional_formatting.add(f"J5:J{end_row}", CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor=LIGHT_AMBER)))
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 10, "B": 28, "C": 52, "D": 11, "E": 14, "F": 12, "G": 14, "H": 14, "I": 12, "J": 11})


def monday(value: date) -> date:
    return value - timedelta(days=value.weekday())


def add_gantt(wb: Workbook, schedule):
    ws = wb.create_sheet("阶段甘特图")
    first_week = monday(schedule[0][1])
    last_week = monday(schedule[-1][2])
    weeks = []
    current = first_week
    while current <= last_week:
        weeks.append(current)
        current += timedelta(days=7)

    end_column = 4 + len(weeks)
    style_title(ws, "任务引擎阶段甘特图", "按周展示单人标准人力计划；色块表示该阶段在当周有计划任务。", end_column)
    headers = ["阶段", "阶段名称", "计划开始", "计划结束"] + [week for week in weeks]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
        if col >= 5:
            ws.cell(4, col).number_format = "m/d"
    style_header(ws, 4, 1, len(headers))

    palette = ["5B9BD5", "70AD47", "ED7D31", "A5A5A5", "4472C4", "FFC000", "2A7F77", "8064A2"]
    for row, ((code, name, _), color) in enumerate(zip(PHASES, palette), 5):
        phase_rows = [(task, start, end) for task, start, end in schedule if task[1] == code]
        phase_start = min(start for _, start, _ in phase_rows)
        phase_end = max(end for _, _, end in phase_rows)
        ws.cell(row, 1, code)
        ws.cell(row, 2, name)
        ws.cell(row, 3, phase_start)
        ws.cell(row, 4, phase_end)
        ws.cell(row, 3).number_format = "yyyy-mm-dd"
        ws.cell(row, 4).number_format = "yyyy-mm-dd"
        for index, week in enumerate(weeks, 5):
            week_end = week + timedelta(days=6)
            if phase_start <= week_end and phase_end >= week:
                ws.cell(row, index, "■")
                ws.cell(row, index).fill = PatternFill("solid", fgColor=color)
                ws.cell(row, index).font = Font(color=WHITE, bold=True)
            else:
                ws.cell(row, index, "")
                ws.cell(row, index).fill = PatternFill("solid", fgColor="F8FBFD")
            ws.cell(row, index).alignment = Alignment(horizontal="center", vertical="center")

    end_row = 4 + len(PHASES)
    style_body(ws, 5, end_row, 1, 4)
    for row in range(5, end_row + 1):
        for col in range(1, end_column + 1):
            ws.cell(row, col).border = border
        ws.row_dimensions[row].height = 28
    for col in range(5, end_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 6.5
    set_widths(ws, {"A": 9, "B": 30, "C": 13, "D": 13})
    ws.freeze_panes = "E5"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_future_plan(wb: Workbook, start_date: date):
    ws = wb.create_sheet("后续增强")
    style_title(ws, "任务引擎后续增强计划", "这些任务用于走向生产化；只有 N1、N2 阻塞第一版 UI，其余可与 UI 开发穿插。", 12)
    headers = ["编号", "工作项", "说明", "难度", "估算工时(h)", "估算人日", "建议开始", "建议结束", "与 UI 关系", "当前状态", "风险/条件", "优先级建议"]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    style_header(ws, 4, 1, len(headers))

    current = start_date
    used_half_days = 0
    for row, task in enumerate(FUTURE_TASKS, 5):
        code, name, desc, hours, difficulty, relation, status = task
        slots = hours // 4
        task_start = current
        remaining = slots
        task_end = current
        while remaining > 0:
            available = 2 - used_half_days
            consumed = min(available, remaining)
            remaining -= consumed
            used_half_days += consumed
            task_end = current
            if used_half_days == 2 and remaining > 0:
                current = next_business_day(current)
                used_half_days = 0
        if used_half_days == 2:
            current = next_business_day(current)
            used_half_days = 0
        risk = "需要真实 DMM/CAN 与厂商驱动" if code == "N7" else ("需先冻结公开 API" if code == "N8" else "-")
        priority = "立即" if code in ("N1", "N2") else ("高" if code in ("N3", "N6", "N7") else "中")
        values = [code, name, desc, difficulty, hours, hours / HOURS_PER_DAY, task_start, task_end, relation, status, risk, priority]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)

    end_row = 4 + len(FUTURE_TASKS)
    style_body(ws, 5, end_row, 1, len(headers))
    for row in range(5, end_row + 1):
        ws.cell(row, 6).number_format = "0.0"
        ws.cell(row, 7).number_format = "yyyy-mm-dd"
        ws.cell(row, 8).number_format = "yyyy-mm-dd"
        ws.row_dimensions[row].height = 54
        for col in (1, 4, 5, 6, 7, 8, 9, 10, 12):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    total_row = end_row + 2
    ws.cell(total_row, 1, "合计")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(total_row, 5, f"=SUM(E5:E{end_row})")
    ws.cell(total_row, 6, f"=SUM(F5:F{end_row})")
    for col in range(1, 13):
        ws.cell(total_row, col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(total_row, col).font = Font(name="Microsoft YaHei", bold=True, color=WHITE)
        ws.cell(total_row, col).border = border
        ws.cell(total_row, col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(total_row, 6).number_format = "0.0"

    ws.conditional_formatting.add(f"I5:I{end_row}", FormulaRule(formula=["LEFT(I5,2)=\"阻塞\""], fill=PatternFill("solid", fgColor=LIGHT_RED)))
    ws.conditional_formatting.add(f"J5:J{end_row}", FormulaRule(formula=["J5=\"受阻于设备条件\""], fill=PatternFill("solid", fgColor=LIGHT_AMBER)))
    ws.auto_filter.ref = f"A4:L{end_row}"
    ws.freeze_panes = "C5"
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 9, "B": 30, "C": 55, "D": 10, "E": 13, "F": 12, "G": 13, "H": 13, "I": 18, "J": 16, "K": 30, "L": 13})


def add_milestones(wb: Workbook, schedule):
    ws = wb.create_sheet("里程碑")
    style_title(ws, "任务引擎里程碑与验收口径", "里程碑以可验证的能力为完成标准。", 7)
    headers = ["里程碑", "名称", "完成标志", "预期结果", "计划完成", "风险提示", "状态"]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col, value)
    style_header(ws, 4, 1, len(headers))
    phase_end = {}
    for code, _, _ in PHASES:
        phase_end[code] = max(end for task, _, end in schedule if task[1] == code)
    ui_start = next_business_day(schedule[-1][2])
    rows = [
        ("M1", "调度内核可运行", "多 UUT、资源、Barrier、Retry、Cleanup、Stop 通过自动化测试", "形成稳定的图调度主链", phase_end["P2"], "Checkpoint 恢复暂不纳入首轮", "未开始"),
        ("M2", "JSON 流程可配置", "Sequence JSON 可编译为不可变 Plan，并提供字段路径错误", "换测试流程无需修改调度代码", phase_end["P3"], "首轮不包含可视化编辑器", "未开始"),
        ("M3", "业务模块可插拔", "Action 可调用内置模块、外部 exe、Python 和符合 ABI 的 DLL", "业务逻辑与调度器解耦", phase_end["P5"], "任意厂商 DLL 仍需要 Adapter", "未开始"),
        ("M4", "动态数据可驱动", "For 循环和运行期变量可驱动每轮不同模块输入", "Plan 保持不可变，循环由调度控制", phase_end["P6"], "复杂 While/条件分支后续扩展", "未开始"),
        ("M5", "测量结果可消费", "测量值、单位、限值、判定、attempt 和 iteration 可进入报告", "CLI/UI 通过只读 DTO 消费结果", max(phase_end["P4"], phase_end["P6"]), "正式报告文件格式后续设计", "未开始"),
        ("M6", "设备连接可跨 Step", "DMM1/CAN1 session 由工位运行时持有，Persistent Host 保持状态", "业务模块不直接管理驱动对象", phase_end["P7"], "并发 Transport 安全需压测", "未开始"),
        ("M7", "DMM/CAN 方向验证", "Fake DMM/CAN 完成 Connect/Configure/Read/Disconnect 端到端", "验证适配层设计可落地", phase_end["P7"], "首轮不连接真实硬件", "未开始"),
        ("M8", "进入 UI 阶段", "Engine/UI 独立工程，ViewModel 只暴露命令和只读 DTO", "开始最小 Runner UI", ui_start, "先做 Runner，不先做完整编辑器", "未开始"),
    ]
    for row, values in enumerate(rows, 5):
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
    end_row = 4 + len(rows)
    style_body(ws, 5, end_row, 1, len(headers))
    for row in range(5, end_row + 1):
        ws.row_dimensions[row].height = 55
        ws.cell(row, 5).number_format = "yyyy-mm-dd"
        for col in (1, 5, 7):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.conditional_formatting.add(f"G5:G{end_row}", FormulaRule(formula=["G5=\"未开始\""], fill=PatternFill("solid", fgColor=LIGHT_AMBER)))
    ws.auto_filter.ref = f"A4:G{end_row}"
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 11, "B": 26, "C": 48, "D": 38, "E": 16, "F": 35, "G": 12})


def main():
    schedule = build_schedule(TASKS)
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    add_overview(wb, schedule)
    add_task_plan(wb, schedule)
    add_phase_summary(wb, schedule)
    add_gantt(wb, schedule)
    add_future_plan(wb, next_business_day(schedule[-1][2]))
    add_milestones(wb, schedule)

    wb.properties.title = "PicoATE 任务引擎研发计划与工作量评估"
    wb.properties.subject = "任务引擎 WBS、工时、排期、里程碑和后续增强计划"
    wb.properties.creator = "PicoATE Project"
    wb.properties.description = "按单人标准工时建立的任务引擎研发计划，包含 WBS、排期、里程碑和后续增强。"
    wb.save(OUTPUT)

    # 重新加载一次，确保生成文件结构有效。
    checked = load_workbook(OUTPUT, read_only=True, data_only=False)
    expected = {"项目概览", "任务计划", "阶段汇总", "阶段甘特图", "后续增强", "里程碑"}
    if set(checked.sheetnames) != expected:
        raise RuntimeError(f"Unexpected worksheets: {checked.sheetnames}")
    checked.close()
    print(OUTPUT)
    print(f"tasks={len(TASKS)}, hours={sum(item[4] for item in TASKS)}, person_days={sum(item[4] for item in TASKS) / HOURS_PER_DAY:.1f}")
    print(f"baseline={schedule[0][1]}..{schedule[-1][2]}")


if __name__ == "__main__":
    main()
